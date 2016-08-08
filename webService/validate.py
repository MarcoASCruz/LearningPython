from flask import Flask, request, jsonify
from json import JSONEncoder
import json

import mysql.connector

import statistics

from datetime import timedelta
import datetime

from sklearn.neighbors import KNeighborsClassifier
from sklearn.neighbors.nearest_centroid import NearestCentroid
from sklearn import tree
from sklearn.cross_validation import StratifiedKFold
from sklearn import svm
from sklearn.naive_bayes import GaussianNB
from sklearn.ensemble import RandomForestClassifier
#from sklearn.neural_network import MLPClassifier

from sklearn.metrics import accuracy_score
import math
import numpy

from openpyxl import Workbook

import matplotlib.pyplot as plt
gLatencies = dict();


app = Flask(__name__)

@app.route('/setSample', methods=['POST'])
def setSample():
	result = "OK"
	models = json.loads(request.form['model'], object_hook = keystrokeDecoder);
	idUser = request.form['idUser'];
	insertKeystrokeDataInDB(idUser, models)
	return result;
	
@app.route('/validate', methods=['POST'])
def validate():
	models = json.loads(request.form['model'], object_hook = keystrokeDecoder);
	idUser = request.form['idUser'];
	insertKeystrokeDataInDB(idUser, models, True);
	# validator = ValidateKeys(selectOriginalModelsFromDB(idUser), selectFakeModelsFromDB(idUser));
	# validated = validator.validate(models);
	#return Encoder().encode(validated);
	return "OK";
	
def insertKeystrokeDataInDB(idUser, models, isTest = False):
	insertString = "INSERT INTO `mdl_user";
	insertKeystroke = "INSERT INTO mdl_user";
	if isTest == True:
		testePrefix = "_test";
		insertString = insertString + testePrefix; #
		insertKeystroke = insertKeystroke + testePrefix;
	insertString = (insertString + "_key_string`(`id_user`, `string`) VALUES (%s, %s)");
	insertKeystroke = (insertKeystroke + "_keystroke (key_code, time, action, id_string) "
				   "VALUES (%(key_code)s, %(time)s, %(action)s, %(id_string)s)");
	
	for model in models: # for each string
		string = createStringFromKeys(model); # for improve the performance I can update the string after the for bellow
		idString = executeSqlInDB(insertString, (idUser, string));
		
		for key in model:
			executeSqlInDB(insertKeystroke, getModelQuery(key, idString))
		
		insertNormalizedModelInDB(idUser, idString, model, isTest);

def createStringFromKeys(keys):
	string = "";
	for key in keys:
		string = string + '{}_'.format(key.keyCode);
	return string[: len(string) - 1];

	
@app.route('/recalculate', methods=['POST', 'GET'])
def recalculate():
	#updating datas from inicial datas
	deleteNormalizedDatas();
	for userString in selectUserStrings():
		keys = selectKeystrokes(userString.idString);
		insertNormalizedModelInDB(userString.idUser, userString.idString, keys);
	
	
	global gLatencies;
	gLatencies = dict();
	#updating datas from posts
	deleteNormalizedDatas(True);
	for userString in selectUserStrings(True):
		keys = selectKeystrokes(userString.idString, True);
		insertNormalizedModelInDB(userString.idUser, userString.idString, keys, True);
	
	plt.bar(gLatencies.keys(), gLatencies.values())
	plt.show()
	
	return "OK";

def deleteNormalizedDatas(isTest = False):
	sql = "DELETE FROM `mdl_user#isTest_keystroke_normalized` WHERE 1 = %(trueValue)s";
	sql = replaceIfIsTest(sql, isTest);
	executeSqlInDB(sql, {'trueValue': 1});


	
def selectUserStrings(isTest = False):
	models = []
	try:
		cnx = mysql.connector.connect(user='root', database='moodle')
		cursor = cnx.cursor();
		
		sql = replaceIfIsTest("SELECT `id_user`, `id` FROM `mdl_user#isTest_key_string` ORDER BY `id_user`", isTest);
		
		cursor.execute(sql);
		for (id_user, id) in cursor:
			models.append(UserString(id_user, id));

		cursor.close();
		cnx.close();
	except mysql.connector.Error as err:
		print(err.msg);
	return models;	

def selectKeystrokes(idString, isTest = False):
	models = []
	try:
		cnx = mysql.connector.connect(user='root', database='moodle')
		cursor = cnx.cursor();
		
		sql = replaceIfIsTest("SELECT key_code, time, action FROM `mdl_user#isTest_keystroke` WHERE `id_string` = %(idString)s  ORDER BY time", isTest);
		
		cursor.execute(sql, {'idString': idString});
		for (key_code, time, action) in cursor:
			models.append(KeystrokeModel(key_code, time, action));

		cursor.close();
		cnx.close();
	except mysql.connector.Error as err:
		print(err.msg);
	return models;
	
def insertNormalizedModelInDB(idUser, idString, keystroke, isTest = False):
	insertNormalizedRecord = replaceIfIsTest("INSERT INTO `mdl_user#isTest_keystroke_normalized`(`id_user`, `id_string`) VALUES (%s, %s)", isTest);
	updateNormalizedRecord = replaceIfIsTest("UPDATE `mdl_user#isTest_keystroke_normalized` ", isTest);
	
		
	executeSqlInDB(insertNormalizedRecord, (idUser, idString));
	
	keyDimensionsExtractor = KeystrokeDimensionsExtractor(keystroke);
	
	#extracting dimensions
	timePressed = keyDimensionsExtractor.getTimePressed();
	#geting avarage and standardDeviation
	timePressedAverage = statistics.mean(timePressed);
	timePressedstandardDeviation = statistics.pstdev(timePressed);
	
	latencies = keyDimensionsExtractor.getLatencies();
	latenciesAverage = statistics.mean(latencies);
	latenciesStandardDeviation = statistics.pstdev(latencies);
	
	dbModel = {
		'id_user': idUser,
		'id_string': idString,
		'press_average': timePressedAverage,
		'latency_avarage': latenciesAverage,
		'press_standard_deviation': timePressedstandardDeviation,
		'latency_standard_deviation': latenciesStandardDeviation,
	}
	
	#update in table created before
	updateNormalizedRecord = updateNormalizedRecord + (" SET `press_average`= %(press_average)s,`latency_avarage`= %(latency_avarage)s, `press_standard_deviation`= %(press_standard_deviation)s,`latency_standard_deviation`= %(latency_standard_deviation)s " 
		" WHERE `id_user`= %(id_user)s AND `id_string`= %(id_string)s");
	executeSqlInDB(updateNormalizedRecord, dbModel);
	
def replaceIfIsTest(string, isTest = False):
	if isTest == True:
		string = string.replace("#isTest", "_test");
	else:
		string = string.replace("#isTest", "");
	return string;
	
def executeSqlInDB(query, queryModel):	
	try:
		cnx = mysql.connector.connect(user='root', database='moodle');
		cursor = cnx.cursor();
		cursor.execute(query, queryModel);
		cnx.commit();
		cursor.close();
		cnx.close();
		return cursor.lastrowid;
	except mysql.connector.Error as err:
		print(err.msg);
		exit(1);

def getModelQuery(keyModel, idString):
	data_model_keystroke = {
		 'key_code': keyModel.keyCode,
		  'time': keyModel.time,
		  'action': keyModel.action,
		  'id_string': idString,
		}
	return data_model_keystroke;

def getDate(dateMilliseconds): 
	timeMilliseconds = timedelta(milliseconds = dateMilliseconds);
	initialDate = datetime.datetime(1970,1,1); #dates in milliseconds start on (1970,1,1)
	return initialDate + timeMilliseconds;
def convertToMilliseconds(microseconds):
	return microseconds / 1000;

	
@app.route('/test', methods=['POST', 'GET'])
def test():
	#model = json.loads(request.form['model'], object_hook = keystrokeDecoder);
	#idUser = 21;
	validateKeys = ValidateKeys();
	#validateKeys.predict(idUser, model);
	validateKeys.predictFromDB();
	return "OK";
	
	
	
	
	
class KeystrokeModel:
	def __init__(self, keyCode, time, action=None):
		self.keyCode = keyCode;
		self.time = time;
		self.action = action;
class KeysNormalizedModel:
	def __init__(self, pressAverage, latencyAverage, pressStandardDeviation, latencyStandardDeviation):
		self.pressAverage = pressAverage;
		self.latencyAverage = latencyAverage;
		self.pressStandardDeviation = pressStandardDeviation;
		self.latencyStandardDeviation = latencyStandardDeviation;
class AlgorithmResultsModel:
	def __init__(self, algorithm, average, standardDeviation):
		self.algorithm = algorithm;
		self.average = average;
		self.standardDeviation = standardDeviation;
class AlgorithmPartialResultsModel:
	def __init__(self, algorithmName):
		self.algorithmName = algorithmName;
		self.accuracySum = 0;
		self.partialAccuracies = [];
class UserString:
	def __init__(self, idUser, idString):
		self.idUser = idUser;
		self.idString = idString;	
		
		
def keystrokeDecoder(obj):
	return KeystrokeModel(obj['keyCode'], obj['time'], obj['action']);

class Encoder(JSONEncoder):
    def default(self, o):
        return o.__dict__  

		
		
class KeystrokeDimensionsExtractor:
	def __init__(self, keys):
		self.keys = keys;
	
	def getTimePressed(self):
		result = [];
		
		for i in range(0, len(self.keys), 1):
			if self.keys[i].action == "DOWN" and self.keyIsValid(self.keys[i].keyCode):
				keyUp = self.findKeyUp(self.keys, i+1, self.keys[i]);
				timePressed = getDate(keyUp.time) - getDate(self.keys[i].time);
				result.append(convertToMilliseconds(timePressed.microseconds));
				
		return result;	
	
	def getLatencies(self):
		result = [];
		
		global gLatencies;
		
		i = 0;
		while i < len(self.keys): #actually, the break will decide when stop
			nextPosition = i + 1;
			keyUp = self.findKeyUp(self.keys, nextPosition, self.keys[i]);
			nextKeyDownPosition = self.findNextKeyDownPosition(self.keys, nextPosition, self.keys[i]);
			if nextKeyDownPosition != None:
				#if (self.keyIsValid(keyUp.keyCode) and self.keyIsValid(self.keys[nextKeyDownPosition].keyCode)):
				#if self.isDigraph(keyUp.keyCode, self.keys[nextKeyDownPosition].keyCode):
				latency = self.calcLatency(keyUp, self.keys[nextKeyDownPosition]);
				#if latency < 600:
				if latency in gLatencies:
					gLatencies[latency] += 1;
				else:
					gLatencies[latency] = 1;
				result.append(latency); 
				i = nextKeyDownPosition;
			else:
				break;

		if(len(result) == 0):
			result.append(0);
		
		
		return result;
		
	def findKeyUp(self, keys, startPosition, keyDown):
		keyUp = None
		for i in range(startPosition, len(keys), 1):
			if keys[i].action == "UP" and (keys[i].keyCode == keyDown.keyCode):
				keyUp = keys[i]
				break
		return keyUp;
	
	def findNextKeyDownPosition(self, keys, startPosition, keyDown):
		nextKeyDownPosition = None;
		for i in range(startPosition, len(keys), 1):
			if keys[i].action == "DOWN":
				nextKeyDownPosition = i;
				break;
		return nextKeyDownPosition;
	
	def keyIsValid(self, keyCode):
		result = False;
		if keyCode >= 48 and keyCode <= 90:
			result = True
		return result;
	
	def isDigraph(self, firstKeyCode, secondKeyCode):
		result = False;
		if (firstKeyCode == 76 and secondKeyCode == 72): #lh
			result = True
		if (firstKeyCode == 78 and secondKeyCode == 72): #nh
			result = True
		if (firstKeyCode == 67 and secondKeyCode == 72): #ch
			result = True
		if (firstKeyCode == 71 and secondKeyCode == 85): #gu
			result = True
		if (firstKeyCode == 81 and secondKeyCode <= 85): #qu
			result = True
		
		if (firstKeyCode == 65 and secondKeyCode == 77): #am
			result = True
		if (firstKeyCode == 65 and secondKeyCode == 78): #an
			result = True
		if (firstKeyCode == 69 and secondKeyCode == 77): #em
			result = True
		if (firstKeyCode == 69 and secondKeyCode == 78): #en
			result = True
		if (firstKeyCode == 73 and secondKeyCode == 77): #im
			result = True
		if (firstKeyCode == 73 and secondKeyCode == 78): #in
			result = True
		if (firstKeyCode == 79 and secondKeyCode == 77): #om
			result = True
		if (firstKeyCode == 79 and secondKeyCode == 78): #on
			result = True
		if (firstKeyCode == 85 and secondKeyCode == 77): #um
			result = True
		if (firstKeyCode == 85 and secondKeyCode == 78): #un
			result = True
		return result;
	
	def calcLatency(self, keyUp, keyDown):
		latency = None;
		if getDate(keyDown.time) > getDate(keyUp.time):
			latency = getDate(keyDown.time) - getDate(keyUp.time);
			latency = convertToMilliseconds(latency.microseconds);
		else:
			latency = getDate(keyUp.time) - getDate(keyDown.time);
			latency = -convertToMilliseconds(latency.microseconds);
		
		return latency;
		
		
class ValidateKeys:
	#def __init__(self):
	
	def predictFromDB(self):
		users = self.selectUsers();
		print(users);
		firstRow = True;
		wb = Workbook()
		# grab the active worksheet
		ws = wb.active
		for idUser in users:
			originalNormalizedModels = self.selectOriginalModelsFromDB(idUser);
			fakeNormalizedModels = self.selectFakeModelsFromDB(idUser);
			
			fakeTestGroup = self.selectFakeTestGroup(idUser); 
			originalTestGroup = self.selectOriginalTestGroup(idUser);
			
			originalTrainingAndTestGroup = numpy.concatenate((originalNormalizedModels, originalTestGroup), axis=0);
			fakeTrainingAndTestGroup = numpy.concatenate((fakeNormalizedModels, fakeTestGroup), axis=0);
			
			validator = ValidateSet(originalTrainingAndTestGroup, fakeTrainingAndTestGroup);
			#validator = ValidateSet(originalNormalizedModels, fakeNormalizedModels);
			results = validator.compareAlgorithmsUsingCrossValidation();
			
			#building the header
			if (firstRow == True):
				ws['A2'] = "Users"
				for i in range(0, len(results), 1):
					nextColumn = (i*2);
					ws.cell(row = 1, column = 2 + nextColumn).value = results[i].algorithm;
					ws.cell(row = 2, column = 2 + nextColumn).value = "Average"
					ws.cell(row = 2, column = 3 + nextColumn).value = "Standard Deviation"
				firstRow = False;
			
			#building the body
			rowInXls = []
			rowInXls.append(idUser);
			for result in results:
				rowInXls.append(result.average);
				rowInXls.append(result.standardDeviation);
			
			ws.append(rowInXls);
				
		wb.save("sample.xlsx");
		return "OK" 
		
	def selectUsers(self):		
		result = [];
		try:
			cnx = mysql.connector.connect(user='root', database='moodle')
			cursor = cnx.cursor();
			
			query = ("SELECT `id_user` FROM `mdl_user_key_string` GROUP BY `id_user`");
			
			cursor.execute(query);
			for id_user in cursor.fetchall():
				result.append(id_user[0]);

			cursor.close();
			cnx.close();
		except mysql.connector.Error as err:
			print(err.msg);
			
		return result;
	
	def selectOriginalModelsFromDB(self, idUser):		
		query = ("SELECT `press_average`, `latency_avarage`, `press_standard_deviation`, `latency_standard_deviation`"
				 " FROM mdl_user_keystroke_normalized" 
				 " WHERE id_user = %(idUser)s");
		return self.selectModelsFromDB(query, idUser, self.vectorSerialize);
	def selectOriginalTestGroup(self, idUser):		
		query = ("SELECT `press_average`, `latency_avarage`, `press_standard_deviation`, `latency_standard_deviation`"
				 " FROM mdl_user_test_keystroke_normalized" 
				 " WHERE id_user = %(idUser)s");
		return self.selectModelsFromDB(query, idUser, self.vectorSerialize);

	def selectFakeModelsFromDB(self, idUser):
		query = ("SELECT `press_average`, `latency_avarage`, `press_standard_deviation`, `latency_standard_deviation`"
				" FROM mdl_user_keystroke_normalized" 
				" WHERE id_user <> %(idUser)s");
		return self.selectModelsFromDB(query, idUser, self.vectorSerialize);
	def selectFakeTestGroup(self, idUser):
		query = ("SELECT `press_average`, `latency_avarage`, `press_standard_deviation`, `latency_standard_deviation`"
				" FROM mdl_user_test_keystroke_normalized" 
				" WHERE id_user <> %(idUser)s");
		return self.selectModelsFromDB(query, idUser, self.vectorSerialize);
	
	def vectorSerialize(self, pressAverage, latencyAvarage, pressStandardDeviation, latencyStandardDeviation):
		return [pressAverage, latencyAvarage, pressStandardDeviation, latencyStandardDeviation]
	
	def selectModelsFromDB(self, query, idUser, serielizeCallback):
		models = []
		try:
			cnx = mysql.connector.connect(user='root', database='moodle')
			cursor = cnx.cursor();
			
			cursor.execute(query, {'idUser': idUser});
			for (press_average, latency_avarage, press_standard_deviation, latency_standard_deviation) in cursor:
				models.append(serielizeCallback(press_average, latency_avarage, press_standard_deviation, latency_standard_deviation));

			cursor.close();
			cnx.close();
		except mysql.connector.Error as err:
			print(err.msg);
		return models;
	


class ValidateSet:
	def __init__(self, originalSet, fakeSet):
		self.trainingGroup = numpy.concatenate((originalSet, fakeSet), axis=0);
		self.trainingGroupClassification = self.getGroupClassification(originalSet, fakeSet);
		
	def getGroupClassification(self, originalSet, fakeSet):
		result = [];
	
		for key in originalSet:
			result.append("T");
		for key in fakeSet:
			result.append("F");
		
		return result
	
	def compareAlgorithmsUsingCrossValidation(self):
		numberOfSets = 10;
		stratifiedFolds = StratifiedKFold(self.trainingGroupClassification, n_folds = numberOfSets);
		print("            Training  Test  Accuracy");
		index = 1;
		partialAccuracies = [];
		
		partialResultsDecitionTree = AlgorithmPartialResultsModel("Decision Tree");
		partialResultsKnn = AlgorithmPartialResultsModel("K-NN");
		partialResultsKnnCentroid = AlgorithmPartialResultsModel("K-NN Centroid");
		partialResultsSVM = AlgorithmPartialResultsModel("SVM-SVC");
		partialResultsNaiveBayes = AlgorithmPartialResultsModel("Naive Bayes");
		partialResultsRandomForest = AlgorithmPartialResultsModel("Random Forest");
		#partialResultsNeuralNetwork = AlgorithmPartialResultsModel("Neural Network - MPL");
		
		for trainingSetIndex, testSetIndex in stratifiedFolds:
			trainingSetValues = [];
			trainingSetClassification = [];
			testSetValues = [];
			testSetClassification = [];
			
			for i in trainingSetIndex:
				trainingSetValues.append(self.trainingGroup[i]);
				trainingSetClassification.append(self.trainingGroupClassification[i]);
			for i in testSetIndex:
				testSetValues.append(self.trainingGroup[i]);
				testSetClassification.append(self.trainingGroupClassification[i]);
			
			partialResultsDecitionTree = self.updatePartialResults(self.predictUsingDecisionTree, partialResultsDecitionTree, trainingSetValues, trainingSetClassification, testSetValues, testSetClassification, index);
			partialResultsKnn = self.updatePartialResults(self.predictUsingKnn, partialResultsKnn, trainingSetValues, trainingSetClassification, testSetValues, testSetClassification, index);
			partialResultsKnnCentroid = self.updatePartialResults(self.predictUsingKnnCentroid, partialResultsKnnCentroid, trainingSetValues, trainingSetClassification, testSetValues, testSetClassification, index);
			partialResultsSVM = self.updatePartialResults(self.predictUsingSVM, partialResultsSVM, trainingSetValues, trainingSetClassification, testSetValues, testSetClassification, index);
			partialResultsNaiveBayes = self.updatePartialResults(self.predictUsingNaiveBayes, partialResultsNaiveBayes, trainingSetValues, trainingSetClassification, testSetValues, testSetClassification, index);
			partialResultsRandomForest = self.updatePartialResults(self.predictUsingRandomForest, partialResultsRandomForest, trainingSetValues, trainingSetClassification, testSetValues, testSetClassification, index);
			#partialResultsNeuralNetwork = self.updatePartialResults(self.predictUsingNeuralNetwork, partialResultsNeuralNetwork, trainingSetValues, trainingSetClassification, testSetValues, testSetClassification, index);
			
			index += 1;
		return self.createAlgorithmResultsModel([partialResultsKnn, partialResultsKnnCentroid, partialResultsDecitionTree, partialResultsSVM, partialResultsNaiveBayes, partialResultsRandomForest])
	
	def updatePartialResults(self, predictAlgorithm, model, trainingSetValues, trainingSetClassification, testSetValues, testSetClassification, index):
		result = predictAlgorithm(trainingSetValues, trainingSetClassification, testSetValues);
		accuracy = accuracy_score(testSetClassification, result);
		
		testGroupLength = len(testSetValues);
		trainingGroupLength = len(trainingSetValues);
		
		model.accuracySum += (accuracy * testGroupLength);
		model.partialAccuracies.append([accuracy, testGroupLength]);
		print("rodada:{:2}".format(index), "{0:7}".format(trainingGroupLength), "{0:7}".format(testGroupLength), "{0:9.2%}".format(accuracy));
		return model;
	
	def createAlgorithmResultsModel(self, models):
		result = [];
		for model in models:
			totalAverage = model.accuracySum/len(self.trainingGroup);
			totalStandardDeviation = self.calcStandardDeviation(model.partialAccuracies, totalAverage);
			result.append(AlgorithmResultsModel(model.algorithmName, "{:9.2%}".format(totalAverage), "{:9.2%}".format(totalStandardDeviation)))
		return result;
		
	def predictUsingKnn(self, trainingGroup, trainingGroupClassification, testDatas):
		neighClassifier = KNeighborsClassifier(n_neighbors = 5, metric="euclidean", weights="distance");
		return self.predict(neighClassifier, trainingGroup, trainingGroupClassification, testDatas);
	
	def predictUsingKnnCentroid(self, trainingGroup, trainingGroupClassification, testDatas):
		neighClassifier = NearestCentroid();
		return self.predict(neighClassifier, trainingGroup, trainingGroupClassification, testDatas);
	
	def predictUsingDecisionTree(self, trainingGroup, trainingGroupClassification, testDatas):
		treeClassifier = tree.DecisionTreeClassifier();
		return self.predict(treeClassifier, trainingGroup, trainingGroupClassification, testDatas);
	
	def predictUsingSVM(self, trainingGroup, trainingGroupClassification, testDatas):
		svmClassifier = svm.SVC();
		return self.predict(svmClassifier, trainingGroup, trainingGroupClassification, testDatas);
	
	def predictUsingNaiveBayes(self, trainingGroup, trainingGroupClassification, testDatas):
		naiveBayesClassifier = GaussianNB();
		return self.predict(naiveBayesClassifier, trainingGroup, trainingGroupClassification, testDatas);
	
	def predictUsingRandomForest(self, trainingGroup, trainingGroupClassification, testDatas):
		randomForestClassifier = RandomForestClassifier();
		return self.predict(randomForestClassifier, trainingGroup, trainingGroupClassification, testDatas);
	
	def predictUsingNeuralNetwork(self, trainingGroup, trainingGroupClassification, testDatas):
		neuralNetworkClassifier = MLPClassifier(algorithm='l-bfgs', alpha=1e-5, hidden_layer_sizes=(5, 2), random_state=1);
		return self.predict(neuralNetworkClassifier, trainingGroup, trainingGroupClassification, testDatas);
	
	def predict(self, classifier, trainingGroup, trainingGroupClassification, testDatas):
		classifier.fit(trainingGroup, trainingGroupClassification);
		result = classifier.predict(testDatas);
		return result;	
	
	def calcStandardDeviation(self, values, average):
		sumValues = 0
		for value in values:
			sumValues += math.pow(value[0] - average, 2) * value[1];
		variance = sumValues / (len(self.trainingGroup) - 1);
		return math.sqrt(variance);
	
	def predictUsingDecisionTreeAndStratified(self):
		NUM_CONJ = 3  #--- número de conjuntos para particionamento dos dados e classes para treinamento

		dados  = self.trainingGroup
		classe = self.trainingGroupClassification
		
		#--- criação dos índices dos elementos de cada folder (conjuntos de treino e teste)
		skf = StratifiedKFold(classe, n_folds=NUM_CONJ)
		rodada = 0
		print("Tam. conjunto: treino  teste")
		for ind_treino, ind_teste in skf:  #--- exibe tamanhos de cada conjunto (treino e teste)
			rodada += 1
			print("--- rodada:{0:2}".format(rodada), "{0:7}".format(len(ind_treino)), "{0:6}".format(len(ind_teste)))	

		rodada = soma_acuracia = 0
		#--- execução para cada conjunto de dados
		print("Rodada Acurácia")
		for ind_treino, ind_teste in skf:
			#--- criação dos folders
			dados_treino = []
			classe_treino = []
			dados_teste = []
			classe_teste = []
			for i in ind_treino:
				dados_treino.append(dados[i])
				classe_treino.append(classe[i])
			for i in ind_teste:
				dados_teste.append(dados[i])
				classe_teste.append(classe[i])

			#--- classificação    
			clf = tree.DecisionTreeClassifier()  #--- definição do classificador
			clf.fit(dados_treino, classe_treino) #--- treino do classificador do comitê nos dados e classes da partição
			result = clf.predict(dados_teste)    #--- resultado do clsasificador nos dados de teste da partição

			acuracia = accuracy_score(classe_teste, result)    #--- acurácia da rodada
			soma_acuracia += (acuracia * len(ind_teste))       #--- soma para média ponderada
			rodada += 1
			print("{0:4} - ".format(rodada), "{:7.2%}".format(acuracia))

		print("Total:{:9.2%}".format(soma_acuracia/len(dados)))

	
	def createClassification(self, dataLength, dataIsOriginal):
		result = [];
		label = "F";
		if dataIsOriginal == True:
			label = "T";
		for i in range(0, dataLength, 1):
			result.append(label);
		return result;
	

		

	






